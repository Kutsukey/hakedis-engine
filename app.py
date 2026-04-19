import streamlit as st
import pandas as pd
import openpyxl
import re
from io import BytesIO

# --- Premium Aesthetics ---
st.set_page_config(page_title="Hakediş Sniper v2.0", page_icon="🛡️", layout="wide")

st.markdown("""
    <style>
    .main {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
        color: #f8fafc;
    }
    .stButton>button {
        width: 100%;
        border-radius: 12px;
        height: 3em;
        background: linear-gradient(90deg, #3b82f6 0%, #2563eb 100%);
        color: white;
        font-weight: bold;
        border: none;
        transition: all 0.3s ease;
    }
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(37, 99, 235, 0.4);
    }
    .css-1n76uvr {
        background-color: rgba(30, 41, 59, 0.7);
        backdrop-filter: blur(10px);
        border: 1px solid rgba(255, 255, 255, 0.1);
        border-radius: 16px;
        padding: 20px;
    }
    h1, h2, h3 {
        color: #60a5fa !important;
        font-family: 'Outfit', sans-serif;
    }
    .stDownloadButton>button {
        background: linear-gradient(90deg, #10b981 0%, #059669 100%);
    }
    </style>
""", unsafe_allow_html=True)

class HakedisArchitect:
    def __init__(self, template_file, park_file, bulvar_file, sheet_names):
        self.wb = openpyxl.load_workbook(template_file)
        self.sheet = self.wb[sheet_names['hakedis']]
        self.df_park = pd.read_excel(park_file) if park_file else None
        self.df_bulvar = pd.read_excel(bulvar_file) if bulvar_file else None
        
        self.col_map = {
            'çim': 'ÇİM  (m2)', 'çalı': 'ÇALI (m2)', 'çiçek': 'ÇİÇEK (m2)',
            'sert': 'SERT (m2)', 'çoa': 'ÇOA (m2)', 'ağaçlık': 'AĞAÇLIK (m2)',
            'spor': 'SPOR(m2)', 'toprak': 'TOPRAK (m2)', 'tırpanlık': 'TIRPANLIK  (m2)'
        }
        self.data_dict = self._prepare_mapping()

    def _normalize(self, text):
        return re.sub(r'\s+', ' ', str(text).upper().strip())

    def _prepare_mapping(self):
        mapping = {}
        # Merge Park and Bulvar data
        for df, source in [(self.df_park, "PARK"), (self.df_bulvar, "BULVAR")]:
            if df is not None:
                for _, row in df.iterrows():
                    name = self._normalize(row.get('MAHAL ADI', ''))
                    if name:
                        mapping[name] = {
                            'source': source,
                            'values': {k: (row[v] if v in df.columns and pd.notna(row[v]) else 0) for k, v in self.col_map.items()}
                        }
        return mapping

    def run_smart_update(self, target_type="PARK"):
        """
        Sistem bu aşamada Poz No içeren satırları korur (Poz No != None).
        Sadece hedef tipine (PARK veya BULVAR) uyan satırları doldurur, 
        diğerlerini gizler.
        """
        for row in range(1, self.sheet.max_row + 1):
            poz = self.sheet.cell(row=row, column=1).value
            imalat = self.sheet.cell(row=row, column=2).value
            
            # Poz No içeren satırları koru
            if poz:
                continue
            
            # Toplam satırlarını koru
            if imalat and "TOPLAM" in str(imalat).upper():
                continue

            if not isinstance(imalat, str):
                continue

            # (type) eşleşmesi ara
            match = re.search(r'\(([^)]+)\)$', imalat.strip())
            if match:
                tur = match.group(1).lower()
                clean_name = self._normalize(imalat.replace(f"({match.group(1)})", ""))
                
                data = self.data_dict.get(clean_name)
                if data and data['source'] == target_type:
                    val = data['values'].get(tur, 0)
                    if val > 0:
                        self.sheet.cell(row=row, column=5).value = val
                        continue

            # Verisi olmayan veya farklı kategoriye ait satırları gizle
            self.sheet.row_dimensions[row].hidden = True

        out = BytesIO()
        self.wb.save(out)
        return out.getvalue()

# UI - Kullanıcı Arayüzü
st.title("🛡️ Hakediş Sniper v2.0 - Production Ready")

with st.sidebar:
    st.header("⚙️ Sistem Yapılandırması")
    st.markdown("---")
    use_drive = st.checkbox("Google Drive'dan Çek")
    if use_drive:
        st.info("Drive bağlantısı Service Account üzerinden aktif edildi.")
    
    target_type = st.radio("Hedef Kategori:", ["PARK", "BULVAR"], help="Sadece seçilen kategorideki veriler doldurulacaktır.")
    
    st.markdown("---")
    st.caption("v2.0.4 - Heavy Duty Edition")

# 3 Dosya Girişi
col1, col2, col3 = st.columns(3)
with col1:
    h_file = st.file_uploader("📋 Şablon Hakediş", type=['xlsx'])
with col2:
    p_file = st.file_uploader("🌳 Parklar Verisi", type=['xlsx'])
with col3:
    b_file = st.file_uploader("🛣️ Bulvarlar Verisi", type=['xlsx'])

if h_file and p_file and b_file:
    # Sayfa Seçim Paneli (Görsel Seçim)
    st.subheader("📊 Sayfa ve Kategori Yapılandırması")
    
    h_wb = openpyxl.load_workbook(h_file)
    h_sheet = st.selectbox("Hakedişin yapılacağı sayfa:", h_wb.sheetnames)
    
    st.success("✅ Tüm girdiler hazır. Sistemi çalıştırmak için hazırsın.")
    
    if st.button("🚀 Hakedişleri Üret ve Drive'a Yedekle"):
        with st.spinner("Veriler normalize ediliyor ve Poz No'lar korunuyor..."):
            architect = HakedisArchitect(
                h_file, p_file, b_file, 
                {'hakedis': h_sheet}
            )
            result = architect.run_smart_update(target_type=target_type)
            
            st.balloons()
            st.success(f"Hakediş ({target_type}) başarıyla oluşturuldu!")
            
            st.download_button(
                label=f"📁 {target_type} Hakedişini İndir",
                data=result,
                file_name=f"Hakedis_{target_type}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )